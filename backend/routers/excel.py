from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, Dict
import pandas as pd
import sqlite3
import json
import os
import uuid
import tempfile
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from database import (
    DB_PATH, get_customers, get_all_tickets, get_deals,
    get_analytics_overview, get_sentiment_trend, get_emotion_breakdown,
    get_action_distribution, get_hourly_activity,
    create_customer, create_ticket, save_conversation_message
)

router = APIRouter(prefix="/excel", tags=["Excel"])

# Temporary storage for uploaded files
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "tmp")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ── Styling constants ──
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
TITLE_FILL = PatternFill(start_color="0f0f1a", end_color="0f0f1a", fill_type="solid")
GREEN_FILL = PatternFill(start_color="1a3a2a", end_color="1a3a2a", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="3a3a1a", end_color="3a3a1a", fill_type="solid")
RED_FILL = PatternFill(start_color="3a1a1a", end_color="3a1a1a", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="333333"),
    right=Side(style="thin", color="333333"),
    top=Side(style="thin", color="333333"),
    bottom=Side(style="thin", color="333333")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, num_cols, row=1):
    """Apply professional header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _add_title_row(ws, title, num_cols):
    """Insert a branded title row at the top and shift data down."""
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"SentimentAI Export — {title}")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    date_cell.font = Font(name="Calibri", italic=True, color="888888", size=9)


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def _apply_borders(ws):
    """Apply thin borders to all data cells."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def _stream_workbook(wb, filename):
    """Return a StreamingResponse from a workbook."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _detect_sheet_type(columns):
    """Detect the type of data based on column names."""
    cols_lower = [c.lower().strip() for c in columns]
    if any(k in cols_lower for k in ["deal", "value", "stage"]):
        return "deals"
    if any(k in cols_lower for k in ["issue", "priority", "status"]):
        return "tickets"
    if any(k in cols_lower for k in ["message", "sentiment"]):
        return "conversations"
    if any(k in cols_lower for k in ["name", "email", "phone"]):
        return "customers"
    return "unknown"


# ──────────────────────────────────────────────
# UPLOAD & IMPORT
# ──────────────────────────────────────────────

class ColumnMappingRequest(BaseModel):
    file_id: str
    column_mapping: Dict[str, str]

class AutoImportRequest(BaseModel):
    file_id: Optional[str] = None

class SyncRequest(BaseModel):
    google_sheet_url: Optional[str] = None
    file_path: Optional[str] = None


@router.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    """Upload an Excel/CSV file, auto-detect type, return preview."""
    if not file.filename:
        raise HTTPException(400, "No file provided")
    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(400, "Only .xlsx, .xls, and .csv files are supported")

    file_id = str(uuid.uuid4())
    save_path = os.path.join(UPLOAD_DIR, f"{file_id}.{ext}")
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)

    try:
        if ext == "csv":
            df = pd.read_csv(save_path)
        else:
            df = pd.read_excel(save_path)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {str(e)}")

    detected_type = _detect_sheet_type(df.columns.tolist())
    preview = df.head(5).fillna("").to_dict(orient="records")
    mapping = {}
    cols_lower = {c.lower().strip(): c for c in df.columns}
    if detected_type == "customers":
        for field in ["name", "email", "phone", "company"]:
            if field in cols_lower:
                mapping[cols_lower[field]] = field
    elif detected_type == "conversations":
        for field in ["message", "sentiment", "emotion", "action"]:
            if field in cols_lower:
                mapping[cols_lower[field]] = field
    elif detected_type == "tickets":
        for field in ["customer_name", "issue", "priority", "status"]:
            if field in cols_lower:
                mapping[cols_lower[field]] = field
    elif detected_type == "deals":
        for field in ["deal", "title", "value", "stage"]:
            if field in cols_lower:
                mapping[cols_lower[field]] = field

    return {
        "file_id": file_id,
        "filename": file.filename,
        "detected_type": detected_type,
        "total_rows": len(df),
        "columns": df.columns.tolist(),
        "preview": preview,
        "mapping": mapping
    }


def _load_uploaded_file(file_id: str):
    """Load a previously uploaded file by ID."""
    for ext in ("xlsx", "xls", "csv"):
        path = os.path.join(UPLOAD_DIR, f"{file_id}.{ext}")
        if os.path.exists(path):
            if ext == "csv":
                return pd.read_csv(path), path
            return pd.read_excel(path), path
    raise HTTPException(404, "Uploaded file not found. Please re-upload.")


def _log_import(filename, file_type, detected_type, total_rows, imported, skipped, errors, imported_ids):
    """Log an import operation to the import_history table."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO import_history (filename, file_type, detected_type, total_rows, imported, skipped, errors, status, imported_ids, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (filename, file_type, detected_type, total_rows, imported, skipped, errors, "completed", json.dumps(imported_ids), datetime.now().isoformat()))
    import_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return import_id


@router.post("/import/customers")
async def import_customers(req: ColumnMappingRequest):
    """Import customers from an uploaded file with column mapping."""
    df, path = _load_uploaded_file(req.file_id)
    reverse_map = {v: k for k, v in req.column_mapping.items()}
    imported_count = 0
    skipped_count = 0
    error_count = 0
    imported_ids = []

    for _, row in df.iterrows():
        try:
            name = str(row.get(reverse_map.get("name", "name"), "")).strip()
            email = str(row.get(reverse_map.get("email", "email"), "")).strip()
            phone = str(row.get(reverse_map.get("phone", "phone"), "")).strip()
            company = str(row.get(reverse_map.get("company", "company"), "")).strip()
            if not name or name == "nan":
                skipped_count += 1
                continue
            email = email if email and email != "nan" else None
            phone = phone if phone and phone != "nan" else None
            company = company if company and company != "nan" else None
            cid = create_customer(name, email, phone, company, "excel_import", "Imported from Excel")
            if cid:
                imported_ids.append(cid)
                imported_count += 1
            else:
                skipped_count += 1
        except Exception:
            error_count += 1

    _log_import(os.path.basename(path), "customers", "customers", len(df), imported_count, skipped_count, error_count, imported_ids)
    return {"imported": imported_count, "skipped": skipped_count, "errors": error_count}


@router.post("/import/conversations")
async def import_conversations(req: ColumnMappingRequest):
    """Import conversations from an uploaded file."""
    df, path = _load_uploaded_file(req.file_id)
    reverse_map = {v: k for k, v in req.column_mapping.items()}
    imported_count = 0
    analyzed_count = 0
    tickets_created = 0
    imported_ids = []

    try:
        from ai.sentiment_classifier import analyze_sentiment, decide_action
        has_ai = True
    except ImportError:
        has_ai = False

    for _, row in df.iterrows():
        try:
            message = str(row.get(reverse_map.get("message", "message"), "")).strip()
            if not message or message == "nan":
                continue
            session_id = f"excel_import_{uuid.uuid4().hex[:8]}"
            if has_ai:
                result = analyze_sentiment(message)
                score = result["score"]
                emotion = result["emotion"]
                action = decide_action(score)
                analyzed_count += 1
            else:
                score = 0.5
                emotion = "neutral"
                action = "NORMAL"

            save_conversation_message(session_id, "user", message, score, emotion, action)
            imported_ids.append(session_id)
            imported_count += 1

            if action == "ESCALATE":
                create_ticket("Excel Import", message, score)
                tickets_created += 1
        except Exception:
            pass

    _log_import(os.path.basename(path), "conversations", "conversations", len(df), imported_count, 0, 0, imported_ids)
    return {"imported": imported_count, "analyzed": analyzed_count, "tickets_created": tickets_created}


@router.post("/import/tickets")
async def import_tickets(req: ColumnMappingRequest):
    """Import tickets from an uploaded file."""
    df, path = _load_uploaded_file(req.file_id)
    reverse_map = {v: k for k, v in req.column_mapping.items()}
    imported_count = 0
    skipped_count = 0
    imported_ids = []

    for _, row in df.iterrows():
        try:
            customer_name = str(row.get(reverse_map.get("customer_name", "customer_name"), "Unknown")).strip()
            issue = str(row.get(reverse_map.get("issue", "issue"), "")).strip()
            if not issue or issue == "nan":
                skipped_count += 1
                continue
            if customer_name == "nan":
                customer_name = "Unknown"
            ticket = create_ticket(customer_name, issue, 0.3)
            if ticket:
                imported_ids.append(ticket["id"])
                imported_count += 1
        except Exception:
            skipped_count += 1

    _log_import(os.path.basename(path), "tickets", "tickets", len(df), imported_count, skipped_count, 0, imported_ids)
    return {"imported": imported_count, "skipped": skipped_count}


@router.post("/import/auto")
async def auto_import(file: UploadFile = File(...)):
    """Full auto-import: detect type and import everything."""
    upload_result = await upload_excel(file)
    file_id = upload_result["file_id"]
    detected_type = upload_result["detected_type"]
    mapping = upload_result["mapping"]

    if detected_type == "unknown":
        return {"type": "unknown", "imported": 0, "skipped": 0, "errors": 0, "preview": upload_result["preview"],
                "message": "Could not auto-detect file type. Please use manual import."}

    req = ColumnMappingRequest(file_id=file_id, column_mapping=mapping)
    if detected_type == "customers":
        result = await import_customers(req)
    elif detected_type == "conversations":
        result = await import_conversations(req)
    elif detected_type == "tickets":
        result = await import_tickets(req)
    else:
        result = {"imported": 0, "skipped": 0, "errors": 0}

    return {"type": detected_type, **result, "preview": upload_result["preview"]}


# ──────────────────────────────────────────────
# EXPORT
# ──────────────────────────────────────────────

@router.get("/export/customers")
def export_customers():
    """Export all CRM customers to a professionally formatted Excel file."""
    customers = get_customers()
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = ["ID", "Name", "Email", "Phone", "Company", "Status", "Risk Score",
               "Avg Sentiment", "Total Conversations", "Total Tickets", "Last Contact"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for c in customers:
        row_data = [
            c.get("id", ""), c.get("name", ""), c.get("email", ""), c.get("phone", ""),
            c.get("company", ""), c.get("status", ""), c.get("risk_score", 0.5),
            c.get("avg_sentiment", 0.5), c.get("total_conversations", 0),
            c.get("total_tickets", 0), c.get("last_contact", "")
        ]
        ws.append(row_data)
        row_num = ws.max_row
        risk = c.get("risk_score", 0.5) or 0.5
        if risk < 0.3:
            fill = RED_FILL
        elif risk < 0.7:
            fill = YELLOW_FILL
        else:
            fill = GREEN_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill

    # Summary row
    ws.append([])
    ws.append(["TOTAL", len(customers)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _add_title_row(ws, "Customer Report", len(headers))
    _auto_fit_columns(ws)
    _apply_borders(ws)
    ws.freeze_panes = "A4"

    return _stream_workbook(wb, "sentimentai_customers.xlsx")


@router.get("/export/conversations")
def export_conversations():
    """Export all conversations with sentiment scores."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM conversations ORDER BY timestamp DESC')
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Conversations"

    headers = ["ID", "Date", "Session", "Role", "Message", "Sentiment Score", "Emotion", "Action", "Language"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        ws.append([
            r.get("id", ""), r.get("timestamp", ""), r.get("session_id", ""),
            r.get("role", ""), r.get("message", ""), r.get("sentiment_score", 0),
            r.get("emotion", ""), r.get("action", ""), r.get("language", "en")
        ])
        row_num = ws.max_row
        action = r.get("action", "NORMAL")
        if action == "ESCALATE":
            fill = RED_FILL
        elif action == "UPSELL":
            fill = GREEN_FILL
        else:
            fill = PatternFill(start_color="2a2a2a", end_color="2a2a2a", fill_type="solid")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill

    ws.append([])
    ws.append(["TOTAL MESSAGES", len(rows)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _add_title_row(ws, "Conversation Log", len(headers))
    _auto_fit_columns(ws)
    _apply_borders(ws)
    ws.freeze_panes = "A4"

    return _stream_workbook(wb, "sentimentai_conversations.xlsx")


@router.get("/export/tickets")
def export_tickets():
    """Export all tickets."""
    tickets = get_all_tickets()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = ["ID", "Customer", "Issue", "Priority", "Status", "Created", "Resolved", "Time to Resolve"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for t in tickets:
        created = t.get("created_at", "")
        resolved = t.get("resolved_at", "")
        time_to_resolve = ""
        if created and resolved:
            try:
                c_dt = datetime.fromisoformat(created)
                r_dt = datetime.fromisoformat(resolved)
                delta = r_dt - c_dt
                hours = delta.total_seconds() / 3600
                time_to_resolve = f"{hours:.1f} hours"
            except:
                pass
        ws.append([
            t.get("id", ""), t.get("customer_name", ""), t.get("issue", ""),
            t.get("priority", ""), t.get("status", ""), created, resolved, time_to_resolve
        ])
        row_num = ws.max_row
        priority = t.get("priority", "LOW")
        if priority == "HIGH":
            fill = RED_FILL
        elif priority == "MEDIUM":
            fill = YELLOW_FILL
        else:
            fill = GREEN_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill

    ws.append([])
    ws.append(["TOTAL TICKETS", len(tickets)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _add_title_row(ws, "Ticket Report", len(headers))
    _auto_fit_columns(ws)
    _apply_borders(ws)
    ws.freeze_panes = "A4"

    return _stream_workbook(wb, "sentimentai_tickets.xlsx")


@router.get("/export/analytics")
def export_analytics():
    """Export full analytics report with multiple sheets."""
    wb = Workbook()

    # Sheet 1: Overview KPIs
    ws1 = wb.active
    ws1.title = "Overview"
    overview = get_analytics_overview()
    kpis = [
        ["Metric", "Value"],
        ["Total Conversations", overview.get("total_conversations", 0)],
        ["Total Tickets", overview.get("total_tickets", 0)],
        ["Resolution Rate", f"{overview.get('resolution_rate', 0)}%"],
        ["Avg Sentiment Score", overview.get("avg_sentiment_score", 0)],
        ["Escalation Rate", f"{overview.get('escalation_rate', 0)}%"],
        ["Upsell Rate", f"{overview.get('upsell_rate', 0)}%"],
        ["Busiest Hour", f"{overview.get('busiest_hour', 0)}:00"],
    ]
    for row in kpis:
        ws1.append(row)
    _style_header_row(ws1, 2)
    _add_title_row(ws1, "Analytics Overview", 2)
    _auto_fit_columns(ws1)
    _apply_borders(ws1)

    # Sheet 2: Daily sentiment trend
    ws2 = wb.create_sheet("Sentiment Trend")
    trend = get_sentiment_trend()
    ws2.append(["Date", "Avg Score", "Count"])
    _style_header_row(ws2, 3)
    for t in trend:
        ws2.append([t.get("date", ""), t.get("avg_score", 0), t.get("count", 0)])
    _add_title_row(ws2, "Daily Sentiment Trend", 3)
    _auto_fit_columns(ws2)
    _apply_borders(ws2)

    # Sheet 3: Emotion breakdown
    ws3 = wb.create_sheet("Emotion Breakdown")
    emotions = get_emotion_breakdown()
    ws3.append(["Emotion", "Count", "Percentage"])
    _style_header_row(ws3, 3)
    for e in emotions:
        ws3.append([e.get("emotion", ""), e.get("count", 0), f"{e.get('percentage', 0)}%"])
    _add_title_row(ws3, "Emotion Breakdown", 3)
    _auto_fit_columns(ws3)
    _apply_borders(ws3)

    # Sheet 4: Action distribution
    ws4 = wb.create_sheet("Action Distribution")
    actions = get_action_distribution()
    ws4.append(["Action", "Count"])
    _style_header_row(ws4, 2)
    for action_name, count in actions.items():
        ws4.append([action_name, count])
    _add_title_row(ws4, "Action Distribution", 2)
    _auto_fit_columns(ws4)
    _apply_borders(ws4)

    # Sheet 5: Hourly activity
    ws5 = wb.create_sheet("Hourly Activity")
    hourly = get_hourly_activity()
    ws5.append(["Hour", "Count"])
    _style_header_row(ws5, 2)
    for h in hourly:
        ws5.append([f"{h.get('hour', 0)}:00", h.get("count", 0)])
    _add_title_row(ws5, "Hourly Activity", 2)
    _auto_fit_columns(ws5)
    _apply_borders(ws5)

    return _stream_workbook(wb, "sentimentai_analytics.xlsx")


@router.get("/export/full-report")
def export_full_report():
    """Complete company report in Excel with charts."""
    wb = Workbook()

    # ── Customers Sheet ──
    ws_cust = wb.active
    ws_cust.title = "Customers"
    customers = get_customers()
    cust_headers = ["ID", "Name", "Email", "Phone", "Company", "Status", "Risk Score", "Avg Sentiment"]
    ws_cust.append(cust_headers)
    _style_header_row(ws_cust, len(cust_headers))
    for c in customers:
        ws_cust.append([
            c.get("id", ""), c.get("name", ""), c.get("email", ""), c.get("phone", ""),
            c.get("company", ""), c.get("status", ""), c.get("risk_score", 0.5), c.get("avg_sentiment", 0.5)
        ])
    _add_title_row(ws_cust, "Full Report — Customers", len(cust_headers))
    _auto_fit_columns(ws_cust)
    _apply_borders(ws_cust)
    ws_cust.freeze_panes = "A4"

    # ── Tickets Sheet ──
    ws_tick = wb.create_sheet("Tickets")
    tickets = get_all_tickets()
    tick_headers = ["ID", "Customer", "Issue", "Priority", "Status", "Created", "Resolved"]
    ws_tick.append(tick_headers)
    _style_header_row(ws_tick, len(tick_headers))
    for t in tickets:
        ws_tick.append([
            t.get("id", ""), t.get("customer_name", ""), t.get("issue", ""),
            t.get("priority", ""), t.get("status", ""), t.get("created_at", ""), t.get("resolved_at", "")
        ])
    _add_title_row(ws_tick, "Full Report — Tickets", len(tick_headers))
    _auto_fit_columns(ws_tick)
    _apply_borders(ws_tick)
    ws_tick.freeze_panes = "A4"

    # ── Conversations Sheet ──
    ws_conv = wb.create_sheet("Conversations")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM conversations ORDER BY timestamp DESC LIMIT 500')
    convos = [dict(r) for r in cursor.fetchall()]
    conn.close()
    conv_headers = ["ID", "Date", "Message", "Score", "Emotion", "Action", "Language"]
    ws_conv.append(conv_headers)
    _style_header_row(ws_conv, len(conv_headers))
    for r in convos:
        ws_conv.append([
            r.get("id", ""), r.get("timestamp", ""), r.get("message", ""),
            r.get("sentiment_score", 0), r.get("emotion", ""), r.get("action", ""), r.get("language", "en")
        ])
    _add_title_row(ws_conv, "Full Report — Conversations", len(conv_headers))
    _auto_fit_columns(ws_conv)
    _apply_borders(ws_conv)
    ws_conv.freeze_panes = "A4"

    # ── Deals Sheet ──
    ws_deals = wb.create_sheet("Deals")
    deals = get_deals()
    deal_headers = ["ID", "Customer", "Title", "Value", "Stage", "Probability", "Expected Close"]
    ws_deals.append(deal_headers)
    _style_header_row(ws_deals, len(deal_headers))
    for d in deals:
        ws_deals.append([
            d.get("id", ""), d.get("customer_name", ""), d.get("title", ""),
            d.get("value", 0), d.get("stage", ""), d.get("probability", 0), d.get("expected_close", "")
        ])
    _add_title_row(ws_deals, "Full Report — Deals", len(deal_headers))
    _auto_fit_columns(ws_deals)
    _apply_borders(ws_deals)
    ws_deals.freeze_panes = "A4"

    # ── Analytics Sheet with Chart ──
    ws_analytics = wb.create_sheet("Analytics")
    overview = get_analytics_overview()
    trend = get_sentiment_trend()
    ws_analytics.append(["KPI", "Value"])
    _style_header_row(ws_analytics, 2)
    ws_analytics.append(["Total Conversations", overview.get("total_conversations", 0)])
    ws_analytics.append(["Total Tickets", overview.get("total_tickets", 0)])
    ws_analytics.append(["Resolution Rate", f"{overview.get('resolution_rate', 0)}%"])
    ws_analytics.append(["Avg Sentiment", overview.get("avg_sentiment_score", 0)])
    ws_analytics.append([])
    ws_analytics.append(["Date", "Avg Sentiment", "Message Count"])
    trend_start_row = ws_analytics.max_row
    _style_header_row(ws_analytics, 3, row=trend_start_row)
    for t in trend:
        ws_analytics.append([t.get("date", ""), t.get("avg_score", 0), t.get("count", 0)])

    if len(trend) > 1:
        chart = LineChart()
        chart.title = "Sentiment Trend"
        chart.style = 10
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Date"
        data_ref = Reference(ws_analytics, min_col=2, min_row=trend_start_row, max_row=ws_analytics.max_row)
        cats_ref = Reference(ws_analytics, min_col=1, min_row=trend_start_row + 1, max_row=ws_analytics.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 20
        chart.height = 12
        ws_analytics.add_chart(chart, f"E{trend_start_row}")

    _add_title_row(ws_analytics, "Full Report — Analytics", 3)
    _auto_fit_columns(ws_analytics)
    _apply_borders(ws_analytics)

    return _stream_workbook(wb, "sentimentai_full_report.xlsx")


# ──────────────────────────────────────────────
# SYNC
# ──────────────────────────────────────────────

@router.post("/sync/crm")
def sync_crm(req: SyncRequest):
    """Sync CRM with an Excel file or Google Sheet."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    now = datetime.now().isoformat()

    if req.file_path:
        try:
            if req.file_path.endswith(".csv"):
                df = pd.read_csv(req.file_path)
            else:
                df = pd.read_excel(req.file_path)
        except Exception as e:
            raise HTTPException(400, f"Cannot read file: {str(e)}")
    elif req.google_sheet_url:
        # Store the config for future reference
        cursor.execute('''
            INSERT OR REPLACE INTO excel_sync_settings (id, sync_type, source_url, status, last_synced, created_at)
            VALUES (1, 'google_sheets', ?, 'connected', ?, ?)
        ''', (req.google_sheet_url, now, now))
        conn.commit()
        conn.close()
        return {"status": "connected", "message": "Google Sheets URL saved. Auto-sync will be available in a future update.",
                "added": 0, "updated": 0, "removed": 0}
    else:
        raise HTTPException(400, "Provide either file_path or google_sheet_url")

    added = 0
    updated = 0
    for _, row in df.iterrows():
        name = str(row.get("name", "")).strip()
        email = str(row.get("email", "")).strip()
        if not name or name == "nan":
            continue
        email = email if email and email != "nan" else None
        phone = str(row.get("phone", "")).strip()
        phone = phone if phone and phone != "nan" else None
        company = str(row.get("company", "")).strip()
        company = company if company and company != "nan" else None

        # Check if customer exists
        if email:
            cursor.execute('SELECT id FROM customers WHERE email = ?', (email,))
        else:
            cursor.execute('SELECT id FROM customers WHERE name = ?', (name,))
        existing = cursor.fetchone()

        if existing:
            update_fields = []
            update_vals = []
            if name and name != "nan":
                update_fields.append("name = ?")
                update_vals.append(name)
            if phone:
                update_fields.append("phone = ?")
                update_vals.append(phone)
            if company:
                update_fields.append("company = ?")
                update_vals.append(company)
            update_fields.append("last_contact = ?")
            update_vals.append(now)
            update_vals.append(existing[0])
            if update_fields:
                cursor.execute(f'UPDATE customers SET {", ".join(update_fields)} WHERE id = ?', tuple(update_vals))
                updated += 1
        else:
            cursor.execute('''
                INSERT INTO customers (name, email, phone, company, source, status, created_at, last_contact, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, email, phone, company, "excel_sync", "lead", now, now, "Synced from Excel"))
            added += 1

    # Update sync settings
    cursor.execute('''
        INSERT OR REPLACE INTO excel_sync_settings (id, sync_type, source_path, status, last_synced, created_at)
        VALUES (1, 'local_file', ?, 'connected', ?, ?)
    ''', (req.file_path, now, now))
    conn.commit()
    conn.close()

    return {"added": added, "updated": updated, "removed": 0}


@router.get("/sync/status")
def get_sync_status():
    """Get current sync status."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM excel_sync_settings ORDER BY id DESC LIMIT 1')
    row = cursor.fetchone()
    conn.close()
    if row:
        return dict(row)
    return {"status": "disconnected", "last_synced": None}


# ──────────────────────────────────────────────
# TEMPLATES
# ──────────────────────────────────────────────

@router.get("/template/{template_type}")
def download_template(template_type: str):
    """Download a blank Excel template with formatted headers and example data."""
    wb = Workbook()
    ws = wb.active

    if template_type == "customers":
        ws.title = "Customer Template"
        headers = ["Name", "Email", "Phone", "Company"]
        ws.append(headers)
        ws.append(["John Doe", "john@example.com", "+1234567890", "Acme Corp"])
        _style_header_row(ws, len(headers))

    elif template_type == "tickets":
        ws.title = "Ticket Template"
        headers = ["Customer Name", "Issue", "Priority", "Status"]
        ws.append(headers)
        ws.append(["Jane Smith", "Login page not loading", "HIGH", "OPEN"])
        _style_header_row(ws, len(headers))
        # Priority dropdown
        dv_priority = DataValidation(type="list", formula1='"LOW,MEDIUM,HIGH"', allow_blank=True)
        dv_priority.prompt = "Select priority"
        ws.add_data_validation(dv_priority)
        dv_priority.add(f"C2:C1000")
        # Status dropdown
        dv_status = DataValidation(type="list", formula1='"OPEN,RESOLVED"', allow_blank=True)
        dv_status.prompt = "Select status"
        ws.add_data_validation(dv_status)
        dv_status.add(f"D2:D1000")

    elif template_type == "conversations":
        ws.title = "Conversation Template"
        headers = ["Message", "Sentiment", "Emotion", "Action"]
        ws.append(headers)
        ws.append(["I am having trouble with my order", "0.3", "frustrated", "ESCALATE"])
        _style_header_row(ws, len(headers))
        dv_action = DataValidation(type="list", formula1='"ESCALATE,NORMAL,UPSELL"', allow_blank=True)
        ws.add_data_validation(dv_action)
        dv_action.add(f"D2:D1000")

    elif template_type == "deals":
        ws.title = "Deal Template"
        headers = ["Title", "Value", "Stage", "Probability", "Expected Close"]
        ws.append(headers)
        ws.append(["Enterprise Plan", 50000, "negotiation", 75, "2025-12-31"])
        _style_header_row(ws, len(headers))
        dv_stage = DataValidation(type="list", formula1='"lead,contacted,negotiation,proposal,won,lost"', allow_blank=True)
        ws.add_data_validation(dv_stage)
        dv_stage.add(f"C2:C1000")
    else:
        raise HTTPException(404, f"Template type '{template_type}' not found. Use: customers, tickets, conversations, deals")

    _auto_fit_columns(ws)
    _apply_borders(ws)
    ws.freeze_panes = "A2"

    return _stream_workbook(wb, f"sentimentai_template_{template_type}.xlsx")


# ──────────────────────────────────────────────
# HISTORY & UNDO
# ──────────────────────────────────────────────

@router.get("/history")
def get_import_history():
    """Get all import history."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM import_history ORDER BY created_at DESC')
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]


@router.post("/undo/{import_id}")
def undo_import(import_id: int):
    """Undo a previous import by deleting imported records."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM import_history WHERE id = ?', (import_id,))
    record = cursor.fetchone()
    if not record:
        conn.close()
        raise HTTPException(404, "Import record not found")

    record = dict(record)
    detected_type = record.get("detected_type", "")
    imported_ids = json.loads(record.get("imported_ids", "[]"))
    deleted_count = 0

    if detected_type == "customers" and imported_ids:
        placeholders = ",".join("?" for _ in imported_ids)
        cursor.execute(f'DELETE FROM customers WHERE id IN ({placeholders})', imported_ids)
        deleted_count = cursor.rowcount
    elif detected_type == "tickets" and imported_ids:
        placeholders = ",".join("?" for _ in imported_ids)
        cursor.execute(f'DELETE FROM tickets WHERE id IN ({placeholders})', imported_ids)
        deleted_count = cursor.rowcount
    elif detected_type == "conversations" and imported_ids:
        placeholders = ",".join("?" for _ in imported_ids)
        cursor.execute(f'DELETE FROM conversations WHERE session_id IN ({placeholders})', imported_ids)
        deleted_count = cursor.rowcount

    cursor.execute('UPDATE import_history SET status = ? WHERE id = ?', ("undone", import_id))
    conn.commit()
    conn.close()
    return {"deleted": deleted_count, "status": "undone"}


# ──────────────────────────────────────────────
# STATS (for frontend cards)
# ──────────────────────────────────────────────

@router.get("/stats")
def get_excel_stats():
    """Get row counts for export preview cards."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM customers')
    customers_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM conversations')
    conversations_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM tickets')
    tickets_count = cursor.fetchone()[0]
    cursor.execute('SELECT COUNT(*) FROM deals')
    deals_count = cursor.fetchone()[0]
    conn.close()
    return {
        "customers": customers_count,
        "conversations": conversations_count,
        "tickets": tickets_count,
        "deals": deals_count
    }


@router.get("/export/preview/{export_type}")
def preview_export(export_type: str):
    """Preview first 20 rows of an export."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if export_type == "customers":
        cursor.execute('SELECT id, name, email, phone, company, status, risk_score, avg_sentiment FROM customers ORDER BY created_at DESC LIMIT 20')
    elif export_type == "conversations":
        cursor.execute('SELECT id, timestamp, message, sentiment_score, emotion, action, language FROM conversations ORDER BY timestamp DESC LIMIT 20')
    elif export_type == "tickets":
        cursor.execute('SELECT id, customer_name, issue, priority, status, created_at, resolved_at FROM tickets ORDER BY created_at DESC LIMIT 20')
    elif export_type == "deals":
        cursor.execute('SELECT * FROM deals ORDER BY created_at DESC LIMIT 20')
    else:
        conn.close()
        raise HTTPException(404, f"Export type '{export_type}' not found")

    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return {"type": export_type, "rows": rows, "columns": list(rows[0].keys()) if rows else []}
